import tkinter as tk 
from tkinter import ttk 
import datetime
from openpyxl import Workbook , load_workbook  

# function of the switch mode button 
def switchMode():
    if modeSwitch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

# Function to search contacts
def searchContact():
    keyword = searchEntry.get().lower()

    # Clear existing data in the treeView
    for item in treeView.get_children():
        treeView.delete(item)

    # Load data from the Excel file
    path = "E:/projects/phoneBook/phoneBook/contacts.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    found = None
    # Iterate through rows and insert matching rows into the treeView
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(keyword in str(cell).lower() for cell in row):
            treeView.insert('', tk.END, values=row)
            found   =   keyword

     # Display messages inside searchFrame based on whether the contact was deleted or not
    if found == keyword:
        searchStatusLabel.config(text=f"Contact '{found}' found")
    else:
        searchStatusLabel.config(text=f"Contact '{found}' not found.")


# Function to sort contacts
def sortContacts():
    path = "E:/projects/phoneBook/phoneBook/contacts.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active

    # Get all data from the sheet
    all_data = list(sheet.iter_rows(min_row=2, values_only=True))

    # Sort the data by the first column (Full Name)
    sorted_data = sorted(all_data, key=lambda x: str(x[0]) if x[0] else '')

    # Clear existing data in the treeView
    for item in treeView.get_children():
        treeView.delete(item)

    # Insert the sorted data into the treeView
    for row in sorted_data:
        treeView.insert('', tk.END, values=row)

# Function to delete a contact
def deleteContact():
    contact_to_delete = deleteEntry.get().lower()

    # Load data from the Excel file
    path = "E:/projects/phoneBook/phoneBook/contacts.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active

    # Check if the contact exists
    contact_found = False
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if contact_to_delete in str(row[0]).lower():
            contact_found = True
            sheet.delete_rows(row_idx)
            break

    # Save the changes to the Excel file
    workbook.save(path)

    # Display messages inside deleteFrame based on whether the contact was found or not
    if contact_found:
        deleteStatusLabel.config(text=f"Contact '{contact_to_delete}' deleted successfully.")
    else:
        deleteStatusLabel.config(text=f"Contact '{contact_to_delete}' not found.")

    # Clear the entry widget
    deleteEntry.delete(0, "end")

# loading data into a table in the TreeViewFrame
def loadData():
    path    =   "E:/projects/phoneBook/phoneBook/contacts.xlsx"
    workbook=   load_workbook(path)
    sheet   =   workbook.active

    listValues  =   list(sheet.values)

    for columnName  in  listValues[0]:
        treeView.heading(columnName,    text=columnName)
    
    for valueTuple  in  listValues[1:]:
        treeView.insert('',tk.END,values=valueTuple)

# inserting the new contacts into excel file
def insertContact():
    fullName=   fullNameEntry.get()
    address =   addressEntry.get()
    email   =   emailEntry.get()
    telephone=  telEntry.get()
    mobile  =   mobileEntry.get()

    # insert row into excel sheet
    path    =   "E:/projects/phoneBook/phoneBook/contacts.xlsx" 
    workbook    =   load_workbook(path)
    sheet   =   workbook.active
    rowValues   =   [fullName,address,email,telephone,mobile]
    sheet.append(rowValues)
    workbook.save(path)

    # insert row into treeView
    treeView.insert('',tk.END,values=rowValues)

    # clear the values
    fullNameEntry.delete(0,"end")
    # fullNameEntry.insert(0, "Full Name")
    addressEntry.delete(0,"end")
    # addressEntry.insert(0,"Address")
    emailEntry.delete(0,"end")
    # emailEntry.insert(0, "Email")
    telEntry.delete(0,"end")
    # telEntry.insert(0,"Telephone Number")
    mobileEntry.delete(0,"end")
    # mobileEntry.insert(0,"Mobile Number")


root    =   tk.Tk()
root.title('Phone Book')
style   =   ttk.Style(root)
root.tk.call("source",  "forest-light.tcl")
root.tk.call("source",  "forest-dark.tcl")
style.theme_use("forest-dark")

frame   =   ttk.Frame(root)
frame.pack()
upframe =   ttk.LabelFrame(frame,text="")
upframe.grid(row=0,column=0)

# list frame 
listFrame =   ttk.LabelFrame(upframe,  text="List the contacts")
listFrame.grid(row=0,  column=0,  padx=5, pady=1)

# listing button
listingButton    =   ttk.Button(listFrame, text="Show the Contacts",  command = loadData)
listingButton.grid(row=5,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")

# sort frame 
sortFrame = ttk.LabelFrame(upframe, text="Sorting the contacts")
sortFrame.grid(row=1, column=0, padx=5, pady=1)

sortButton = ttk.Button(sortFrame, text="Sort by  Name", command=sortContacts)
sortButton.grid(row=0, column=0, padx=5, pady=[0,5], sticky="nsew")

# customize frame
customizeFrame =   ttk.LabelFrame(upframe,  text="Customize")
customizeFrame.grid(row=2,  column=0,  padx=5, pady=1)
# switching mode button
modeSwitch  =   ttk.Checkbutton(customizeFrame,    text="Switch mode", style="Switch", command = switchMode)
modeSwitch.grid(row=0,  column=0,   padx=5, pady=10 ,   sticky="nsew")

# insert frame
insertFrame =   ttk.LabelFrame(upframe,   text    =   "insert new contact")
insertFrame.grid(row=0,  column=1,  padx=5,    pady=1)
#full name entry
nameEntryFrame   =   ttk.LabelFrame(insertFrame, text="Full Name")
nameEntryFrame.grid(row=0,column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")
fullNameEntry = ttk.Entry(nameEntryFrame)
fullNameEntry.grid(row=0,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")
#address entry
addressEntryFrame    =   ttk.LabelFrame(insertFrame, text="Address")
addressEntryFrame.grid(row   =   1,column=0,padx=5,    pady=[0,5],   sticky=  "nsew")
addressEntry = ttk.Entry(addressEntryFrame)
addressEntry.grid(row=0,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")
#email entry
emailEntryFrame  =   ttk.LabelFrame(insertFrame, text="Email")
emailEntryFrame.grid(row=2,  column=0,padx=5,    pady=[0,5],   sticky=  "nsew")
emailEntry = ttk.Entry(emailEntryFrame)
emailEntry.grid(row=0,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")
#telephone entry
telEntryFrame   =   ttk.LabelFrame(insertFrame, text="Telephone")
telEntryFrame.grid(row=3,   column=0,padx=5,    pady=[0,5],   sticky=  "nsew")
telEntry = ttk.Entry(telEntryFrame)
telEntry.grid(row=0,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")
#mobile number entry
mobileEntryFrame    =   ttk.LabelFrame(insertFrame, text="Mobile Number")
mobileEntryFrame.grid(row=4,    column=0,padx=5,    pady=[0,5],   sticky=  "nsew")
mobileEntry = ttk.Entry(mobileEntryFrame)
mobileEntry.grid(row=0,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")

# insert button
insertButton    =   ttk.Button(insertFrame, text="Insert",  command = insertContact)
insertButton.grid(row=5,   column=0   ,padx=5,    pady=[0,5],   sticky=  "nsew")


# Search frame
searchFrame =   ttk.LabelFrame(upframe,   text    =   "Search a contact")
searchFrame.grid(row=0,  column=2,  padx=5,    pady=1)

# searh entry frame
searchEntryFrame = ttk.LabelFrame(searchFrame,  text="Enter a name ")
searchEntryFrame.grid(row=0, column=0, padx=5, pady=[0,5], sticky="nsew")
searchEntry = ttk.Entry(searchEntryFrame)
searchEntry.grid(row=0, column=0, padx=5, pady=[0,5], sticky="nsew")

# Button to perform search
searchButton = ttk.Button(searchFrame, text="Search", command=searchContact)
searchButton.grid(row=1, column=0, padx=5, pady=[0,5], sticky="nsew")

# Label to display status messages
searchStatusLabel = ttk.Label(searchFrame, text=" .... ")
searchStatusLabel.grid(row=3, column=0, padx=5, pady=[0,5], sticky="nsew")

# Result treeView for displaying search results
searchResultFrame = ttk.Frame(searchFrame)
searchResultFrame.grid(row=2, column=0, padx=5, pady=[0,5], sticky="nsew")

# delete frame 
deleteFrame = ttk.LabelFrame(upframe, text="Sorting the contacts")
deleteFrame.grid(row=1, column=2, padx=5, pady=1)

# Entry frame for deleteFrame
deleteEntryFrame = ttk.LabelFrame(deleteFrame, text="Enter a contact")
deleteEntryFrame.grid(row=0, column=0, padx=5, pady=[0,5], sticky="nsew")

# Entry widget for deleting a contact
deleteEntry = ttk.Entry(deleteEntryFrame)
deleteEntry.grid(row=0, column=0, padx=5, pady=[0,5], sticky="nsew")

# Button to perform deletion
deleteButton = ttk.Button(deleteFrame, text="Delete", command=deleteContact)
deleteButton.grid(row=1, column=0, padx=5, pady=[0,5], sticky="sew")

# Label to display status messages
deleteStatusLabel = ttk.Label(deleteFrame, text="  ....  ")
deleteStatusLabel.grid(row=2, column=0, padx=5, pady=[0,5], sticky="nsew")

# TreeView frame
treeViewFrame   =   ttk.LabelFrame(upframe,text="Contacts")
treeViewFrame.grid(row=1,column=1,padx=5, pady=[0,5], sticky="nsew")

# scroll Bar
treeScroll  =   ttk.Scrollbar(treeViewFrame)
treeScroll.grid(row=0, column=1, sticky="ns")

cols    =   ("Full Name",   "Address"   ,   "Email" ,   "Telephone" ,   "Mobile Number")
treeView    =   ttk.Treeview(treeViewFrame, show="headings" ,
        yscrollcommand=treeScroll.set,   column= cols,   height =   5)
treeView.column("Full Name" ,  width=150)
treeView.column("Address" , width=150)
treeView.column("Email" ,   width=150)
treeView.column("Telephone" ,   width=100)
treeView.column("Mobile Number" ,   width=100)
treeView.grid(row=0,column=0)
treeScroll.config(command=treeView.yview)

root.mainloop()